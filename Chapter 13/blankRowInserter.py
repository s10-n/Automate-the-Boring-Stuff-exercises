#! python3
# blankRowInserter.py - takes two integers (N and M) and a spreadsheet filename string as command line arguments, and starting at row N, the program inserts M blank rows into the spreadsheet

import openpyxl,os,sys
from openpyxl.utils import get_column_letter
from pathlib import Path

N = int(sys.argv[1])
M = int(sys.argv[2])
filename_string = sys.argv[3]

# read the contents of the spreadsheet
spreadsheet = openpyxl.load_workbook(filename_string)
print('Opened ' + Path(filename_string).name + '.')
sheet = spreadsheet.active

# assign the rows to be divided to variables
first_N_lines = list(sheet.rows)[:N-1]
remaining_lines = list(sheet.rows)[N:]

# create a new spreadsheet 
new_spreadsheet = openpyxl.Workbook()
new_sheet = new_spreadsheet.active

# write the first N-1 lines to the new spreadsheet unchanged
for row in first_N_lines:
    for cell in row:
        new_sheet[cell.coordinate] = cell.value

# move the remaining rows down by M to create the break
for row in remaining_lines:
    for cell in row:
        new_sheet[get_column_letter(cell.column) + str(cell.row + M)] = cell.value

# save the new spreadsheet
new_filename = Path(filename_string).stem + '_copy.xlsx'
new_spreadsheet.save(new_filename)
print(f'Saved as {new_filename}.')

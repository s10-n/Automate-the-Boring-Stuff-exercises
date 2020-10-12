#! python3
# spreadsheetToTextFiles.py - reads in a spreadsheet and inserts its contents into several text files, with one row per line of text

import openpyxl,os
from openpyxl.utils import get_column_letter

# open the spreadsheet containing the text file data
workbook = openpyxl.load_workbook('/home/sean/projects/automate-the-boring-stuff-exercises/Chapter 13/results.xlsx')
sheet = workbook.active

# read the maximum number of columns (which is the number of text files that the program will be creating
number_of_columns = sheet.max_column

for column_index_number in range(0,number_of_columns):

    # open a text file for each column
    filename = f'text_file_{column_index_number + 1}.txt'
    text_file = open(f'/home/sean/projects/automate-the-boring-stuff-exercises/Chapter 13/text-files/{filename}','w')
    print(f'Now writing to {filename}...')

    # get a tuple of all of the cells in the given column
    column = tuple(sheet.columns)[column_index_number]

    # reset the number of lines written
    line_count = 0

    # iterate through each cell
    for cell_index in range(len(column)):
        cell = sheet[get_column_letter(column_index_number+1) + str(cell_index+1)]
        line_text = cell.value

        # if the cell contains a string, write its value to a new line in the text file and increase the line count by 1
        if isinstance(line_text, str):
            text_file.write(line_text)
            line_count += 1
    print(f'Wrote {line_count} lines to {filename}.')

    # close the current text file
    text_file.close()     

#! python3
# multiplicationTable.py - takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet
import openpyxl, sys, os
from openpyxl.styles import Font

# take in a number
N = int(sys.argv[1])

# create a new spreadsheet
table_workbook = openpyxl.Workbook()
table_sheet = table_workbook.create_sheet(index=0, title=f'{N}x{N} Multiplication Table')
bold_font = Font(bold=True)

# fill row 1 and column A in with an iteration of that number from 1-N
for i in range(1,N + 1):
    table_sheet.cell(row=1,column=(i + 1)).font = bold_font
    table_sheet.cell(row=1,column=(i + 1)).value = i
    table_sheet.cell(row=(i + 1),column=1).font = bold_font
    table_sheet.cell(row=(i + 1),column=1).value = i
# for the remaining cells, use the label rows to determine a formula for what they should be
for x in range(2,N + 2):
    for y in range(2,N + 2):
        table_sheet.cell(row=y,column=x).value = (table_sheet.cell(row=y,column=1).value * table_sheet.cell(row=1,column=x).value)
    
table_workbook.save(f'{N}x{N}_Multiplication_Table.xlsx')
print(f'Wrote "{N}x{N}_Multiplication_Table.xlsx" to ' + str(os.getcwd()))

#! python3
# textFilesToSpreadsheet.py - reads in the contents of several text files and inserts their contents into a spreadsheet, with one line of text per row

import openpyxl,os
from pathlib import Path
from openpyxl.utils import get_column_letter

# create a new spreadsheet to print the results to
workbook = openpyxl.Workbook()
sheet = workbook.active

# start writing in the first column of the spreadsheet
active_column_number = 1

# iterate through the files in the given directory
for file in os.listdir('/home/sean/projects/automate-the-boring-stuff-exercises/Chapter 13/text-files'):

    # convert the active column number into an alphabetical index
    active_column = get_column_letter(active_column_number)

    # open the text file and print its lines to a list of lines
    text_file = open('/home/sean/projects/automate-the-boring-stuff-exercises/Chapter 13/text-files/' + file)
    list_of_lines = text_file.readlines()

    # start writing in the first row of the spreadsheet
    active_row = 1

    # write each line to a new row in the current column
    for line in list_of_lines:
        sheet[active_column + str(active_row)] = line

        # increase the row/column index by 1 when done
        active_row += 1        
    active_column_number += 1

# save the workbook to a new spreadsheet
workbook.save('/home/sean/projects/automate-the-boring-stuff-exercises/Chapter 13/results.xlsx')

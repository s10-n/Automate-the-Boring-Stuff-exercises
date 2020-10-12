#! python3
# spreadsheetCellInverter.py - inverts the row and column of the cells in a given spreadsheet

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

spreadsheet_path = input('Enter a path to a .xlsx file: ')

# read the contents of the spreadsheet

filename = Path(spreadsheet_path).name
old_workbook = openpyxl.load_workbook(spreadsheet_path)
print(f'Loaded {filename}.')
old_sheet = old_workbook.active

# determine the max column and row and save all the filled-in cells in the sheet to a variable
last_column = get_column_letter(old_sheet.max_column)
last_row = old_sheet.max_row
active_cells = tuple(old_sheet['A1':last_column + str(last_row)])

# create a new spreadsheet
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# write the active cells to the new sheet, inverting their X and Y coordinates
for row in active_cells:
    for old_cell in row:
        new_cell_row = old_cell.column
        new_cell_column = old_cell.row
        new_cell = new_sheet.cell(row = new_cell_row, column = new_cell_column)
        new_sheet[new_cell.coordinate] = old_cell.value

# save the new sheet
new_filename = Path(filename).stem + '_inverted.xlsx'
new_workbook.save(new_filename)
print(f'Saved as {new_filename}.')
